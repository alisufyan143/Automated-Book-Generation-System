"""
Input Handler module for the Book Generation System.
Reads book data from Excel files.
"""

from pathlib import Path
from typing import List, Dict, Any, Optional
import openpyxl
from config import Config


class InputHandler:
    """Handles reading book data from Excel files."""
    
    REQUIRED_COLUMNS = ["title"]
    OPTIONAL_COLUMNS = ["notes_on_outline_before"]
    
    def __init__(self, file_path: Optional[str] = None):
        """
        Initialize input handler with Excel file path.
        
        Args:
            file_path: Path to Excel file. Defaults to input/books_input.xlsx
        """
        if file_path:
            self.file_path = Path(file_path)
        else:
            self.file_path = Config.INPUT_DIR / "books_input.xlsx"
    
    def read_books(self) -> List[Dict[str, Any]]:
        """
        Read books from Excel file.
        
        Returns:
            List of book dictionaries with title and notes.
        """
        if not self.file_path.exists():
            raise FileNotFoundError(f"Input file not found: {self.file_path}")
        
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook.active
        
        # Get headers from first row
        headers = [cell.value.lower().strip() if cell.value else "" for cell in sheet[1]]
        
        # Validate required columns
        for col in self.REQUIRED_COLUMNS:
            if col not in headers:
                raise ValueError(f"Missing required column: {col}")
        
        # Read data rows
        books = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_data = dict(zip(headers, row))
            
            # Skip empty rows
            if not row_data.get("title"):
                continue
            
            book = {
                "title": str(row_data["title"]).strip(),
                "notes_on_outline_before": row_data.get("notes_on_outline_before", ""),
                "_row_number": row_idx  # Track source row for debugging
            }
            
            # Clean up notes
            if book["notes_on_outline_before"]:
                book["notes_on_outline_before"] = str(book["notes_on_outline_before"]).strip()
            
            books.append(book)
        
        workbook.close()
        return books
    
    def validate_books(self, books: List[Dict[str, Any]]) -> Dict[str, List]:
        """
        Validate books and separate into valid/invalid.
        
        Returns:
            Dict with 'valid' and 'invalid' lists
        """
        valid = []
        invalid = []
        
        for book in books:
            # Check if has pre-outline notes (required per spec)
            if not book.get("notes_on_outline_before"):
                invalid.append({
                    **book,
                    "_error": "Missing notes_on_outline_before - required before generating outline"
                })
            else:
                valid.append(book)
        
        return {"valid": valid, "invalid": invalid}
    
    def get_books_for_processing(self) -> List[Dict[str, Any]]:
        """
        Read and validate books, returning only valid ones.
        
        Returns:
            List of validated books ready for outline generation
        """
        books = self.read_books()
        result = self.validate_books(books)
        
        if result["invalid"]:
            print(f"⚠️  {len(result['invalid'])} book(s) skipped (missing notes_on_outline_before):")
            for book in result["invalid"]:
                print(f"   - Row {book['_row_number']}: {book['title']}")
        
        return result["valid"]


# ==========================================================================
# CREATE SAMPLE INPUT FILE
# ==========================================================================
def create_sample_input():
    """Create a sample Excel input file."""
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Books"
    
    # Headers
    ws["A1"] = "title"
    ws["B1"] = "notes_on_outline_before"
    
    # Sample data
    ws["A2"] = "The Future of Artificial Intelligence"
    ws["B2"] = "Focus on practical applications in healthcare and education. Include ethical considerations. Target audience: general readers with some tech background."
    
    ws["A3"] = "Mastering Python Programming"
    ws["B3"] = "Beginner to intermediate level. Cover data structures, OOP, and web development basics. Include hands-on projects."
    
    ws["A4"] = "Book Without Notes"
    ws["B4"] = ""  # This one will be skipped
    
    # Adjust column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 80
    
    # Save
    output_path = Config.INPUT_DIR / "books_input.xlsx"
    wb.save(output_path)
    print(f"✅ Sample input file created: {output_path}")
    return output_path


# ==========================================================================
# TEST
# ==========================================================================
if __name__ == "__main__":
    print("=" * 60)
    print("INPUT HANDLER TEST")
    print("=" * 60)
    
    # Create sample input if doesn't exist
    sample_path = Config.INPUT_DIR / "books_input.xlsx"
    if not sample_path.exists():
        print("\n📝 Creating sample input file...")
        create_sample_input()
    
    # Test reading
    print("\n📖 Reading books from input file...")
    handler = InputHandler()
    
    try:
        books = handler.get_books_for_processing()
        print(f"\n✅ Found {len(books)} valid book(s) for processing:\n")
        
        for i, book in enumerate(books, 1):
            print(f"{i}. {book['title']}")
            print(f"   Notes: {book['notes_on_outline_before'][:50]}...")
            print()
            
    except Exception as e:
        print(f"❌ Error: {e}")
